from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import List, Optional, Tuple

from more_itertools import partition
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pyramid.httpexceptions import HTTPFound
from pyramid.response import FileResponse, Response
from pyramid.view import view_config, view_defaults
from pyramid_mailer import get_mailer
from pytz import timezone

from zam_repondeur.data_sanitize import get_as_int_or_none
from zam_repondeur.mails import (
    send_coordinator_invitations,
    send_existing_users_invitations,
    send_new_users_invitations,
    sort_emails,
)
from zam_repondeur.message import Message
from zam_repondeur.models import (
    Chambre,
    DBSession,
    Dossier,
    Lecture,
    MissionSenat,
    Team,
    User,
    get_one_or_create,
)
from zam_repondeur.models.events.dossier import (
    ArchiverDossier,
    DossierRetrait,
    InvitationCoordinateurEnvoyee,
    InvitationEnvoyee,
    PasseEnContributeur,
    PasseEnCoordinateur,
)
from zam_repondeur.models.events.dossiers_list import (
    ArchiverDossiersList
)
from zam_repondeur.models.events.lecture import RefreshLecture
from zam_repondeur.resources import DossierResource
from zam_repondeur.services.data import repository
from zam_repondeur.services.fetch.an.amendements import build_pattern
from zam_repondeur.services.fetch.an.dossiers.dossiers_legislatifs import (
    get_url_dossiers,
)
from zam_repondeur.services.fetch.an.parser import get_parseur
from zam_repondeur.services.fetch.articles import get_texte_url_senat
from zam_repondeur.services.fetch.senat.amendements import _build_amendements_url
from zam_repondeur.services.fetch.senat.derouleur import derouleur_urls_and_mission_refs
from zam_repondeur.services.fetch.senat.scraping import build_rss_url
from zam_repondeur.services.import_export.xlsx import (
    DARK_BLUE,
    WHITE,
    get_user_list_workbook,
)
from zam_repondeur.tasks.fetch import fetch_amendements
from zam_repondeur.views.jinja2_filters import enumeration

from .dossier import DossierViewBase


class DossierInviteAction:
    def _find_or_create_users(self, emails: List[str]) -> Tuple[List[User], List[User]]:
        new_users = []
        existing_users = []
        for email in emails:
            user, created = get_one_or_create(User, email=email)
            if created:
                new_users.append(user)
            else:
                existing_users.append(user)
        return new_users, existing_users

    def _identify_members(
        self, users: List[User], team: Team
    ) -> Tuple[List[User], List[User]]:
        not_members, members = partition(team.is_member, users)
        return list(not_members), list(members)


@view_defaults(context=DossierResource, name="invite_contributeur")
class DossierInviteContributeurForm(DossierViewBase, DossierInviteAction):
    @view_config(request_method="POST")
    def post(self) -> Response:

        origin: str = self.request.POST.get("origin")

        # Only consider well formed addresses from allowed domains
        bad_emails, clean_emails = sort_emails(self.request.POST.get("emails"))

        # Create user accounts if needed
        new_users, existing_users = self._find_or_create_users(clean_emails)

        team = self.dossier.team

        # Users that already have a Signale account, but are not yet members
        new_members, existing_members = self._identify_members(existing_users, team)

        users_to_invite = new_users + new_members

        team.add_members(users_to_invite)

        # get commons mail values
        mailer = get_mailer(self.request)
        url_dossier = self.request.resource_url(self.request.context)
        contact_mail_from = self.request.registry.settings["zam.contact_mail_from"]

        invitations_sent = 0
        if new_users:
            invitations_sent += send_new_users_invitations(
                mailer,
                self.request.user,
                new_users,
                contact_mail_from,
                self.dossier.titre,
                url_dossier,
                self.request.registry,
            )
        if new_members:
            invitations_sent += send_existing_users_invitations(
                mailer,
                self.request.user,
                new_members,
                contact_mail_from,
                self.dossier.titre,
                url_dossier,
                self.request.registry,
            )

        for user in users_to_invite:
            InvitationEnvoyee.create(
                dossier=self.dossier, email=user.email, request=self.request
            )

        cls, message = get_message_from_invitation_result(
            invitations_sent, existing_members, bad_emails
        )

        self.request.session.flash(Message(cls=cls, text=message))

        return HTTPFound(location=self.request.resource_url(self.context, origin))


@view_defaults(
    context=DossierResource, name="invite_coordinateur", permission="gestion"
)
class DossierInviteCoordinateurForm(DossierViewBase, DossierInviteAction):
    @view_config(request_method="POST")
    def post(self) -> Response:

        origin: str = self.request.POST.get("origin")

        # Only consider well formed addresses from allowed domains
        bad_emails, clean_emails = sort_emails(self.request.POST.get("emails"))

        # Create user accounts if needed
        new_users, existing_users = self._find_or_create_users(clean_emails)

        team = self.dossier.team

        users_to_invite = new_users + existing_users

        team.add_coordinators(users_to_invite)

        # get commons mail values
        mailer = get_mailer(self.request)
        url_dossier = self.request.resource_url(self.request.context)
        contact_mail_from = self.request.registry.settings["zam.contact_mail_from"]

        invitations_sent = 0
        invitations_sent += send_coordinator_invitations(
            mailer,
            self.request.user,
            users_to_invite,
            contact_mail_from,
            self.dossier.titre,
            url_dossier,
            self.request.registry,
        )

        for user in users_to_invite:
            InvitationCoordinateurEnvoyee.create(
                dossier=self.dossier, email=user.email, request=self.request
            )

        cls, message = get_message_from_invitation_result(
            invitations_sent, [], bad_emails
        )

        self.request.session.flash(Message(cls=cls, text=message))

        return HTTPFound(location=self.request.resource_url(self.context, origin))


@view_defaults(context=DossierResource, name="invite")
class DossierInviteForm(DossierViewBase):
    @view_config(request_method="GET", renderer="dossier_invite.html")
    def get(self) -> dict:
        return {
            "dossier": self.dossier,
            "dossier_resource": self.context,
            "team": self.dossier.team,
            "current_tab": "invite",
        }


@view_defaults(context=DossierResource, permission="gestion")
class DossierGestionExportsXLS(DossierViewBase):
    @view_config(request_method="GET", name="export_team_xls")
    def get_teams(self) -> Response:
        team = self.dossier.team
        wb = get_user_list_workbook(team, self.dossier.slug)

        prefix, suffix = self._get_prefix_suffix_for_file(self.dossier.slug, "Equipe")

        with NamedTemporaryFile(
            prefix=prefix, suffix=suffix, delete=True
        ) as attachment:
            wb.save(attachment.name)
            wb.close()
            response = FileResponse(attachment.name)
            response.content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response.headers[
                "Content-Disposition"
            ] = f"attachment; filename={prefix}{suffix}"
            return response

    @view_config(request_method="GET", name="export_urls_xls")
    def get_links(self) -> Response:
        wb = Workbook()
        ws = wb[wb.sheetnames[0]]
        ws.title = f"Alimentation_données_{self.dossier.slug}"[0:30]
        ws.sheet_properties.tabColor = "1072BA"

        url_data, lectures = get_all_urls(self.dossier)

        # Configuration des infos du dossier
        infos_dossier = {
            "Informations sur le dossier": "",
            "Identifiant AN": self.dossier.an_id if self.dossier.an_id else "",
            "Identifiant Sénat": self.dossier.senat_id if self.dossier.senat_id else "",
            "Slug": self.dossier.slug,
            "Open Data" if self.dossier.an_id else "Flux XML": url_data,
        }
        row: int = 1
        for titre, info in infos_dossier.items():
            ws[f"A{row}"] = titre
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = info
            row += 1

        # Configuration de l'entête pour les lectures
        ws[f"A{row + 2}"] = "Liens de chaque lecture du dossier"
        ws.append(
            (
                "Chambre",
                "Organe",
                "Phase",
                "Num texte",
                "Lien texte",
                "Lien dérouleur",
                "Missions",
                "Lien(s) ordre de discussion",
            )
        )
        ws.merge_cells("H9:I9")
        ws["H10"] = "HTML"
        ws["I10"] = "JSON"
        for column in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.merge_cells(f"{column}9:{column}10")
            ws[f"{column}9"].alignment = Alignment(vertical="center")

        # Insertion des liens pour chaque lecture
        for dict_links in lectures:
            ws.append(
                (
                    dict_links.get("chambre"),
                    dict_links.get("organe"),
                    dict_links.get("phase"),
                    dict_links.get("num_texte"),
                    dict_links.get("link_texte"),
                    dict_links.get("link_derouleur"),
                    dict_links.get("missions"),
                    "\n".join(
                        [f"{link}.html" for link in dict_links.get("links_discuss")]
                    ),
                    "\n".join(
                        [f"{link}.json" for link in dict_links.get("links_discuss")]
                    ),
                )
            )

        # Redimensionnement de toutes les colonnes
        # Et fusion des deux titres principaux de la feuille
        for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[column].width = 15
            ws[f"{column}9"].font = Font(bold=True)
            ws[f"{column}10"].font = Font(bold=True)
        ws.merge_cells("A1:I1")
        ws.merge_cells("A8:I8")
        ws["A1"].font = Font(color=WHITE, bold=True)
        ws["A8"].font = Font(color=WHITE, bold=True)
        ws["A1"].fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)
        ws["A8"].fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)

        prefix, suffix = self._get_prefix_suffix_for_file(
            self.dossier.slug, "Alimentation_donnees"
        )

        with NamedTemporaryFile(
            prefix=prefix, suffix=suffix, delete=True
        ) as attachment:
            wb.save(attachment.name)
            wb.close()
            response = FileResponse(attachment.name)
            response.content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response.headers[
                "Content-Disposition"
            ] = f"attachment; filename={prefix}{suffix}"
            return response

    @staticmethod
    def _get_prefix_suffix_for_file(slug: str, title: str) -> Tuple[str, str]:
        prefix = f"{title}_{slug}"
        now = datetime.now()
        my_tz = timezone("Europe/Paris")
        date_str = now.astimezone(my_tz).strftime("%Y_%m_%d_%H_%M")
        return (f"{prefix}-{date_str}", ".xlsx")


@view_defaults(context=DossierResource, name="gestion", permission="gestion")
class DossierGestionForm(DossierViewBase):
    @view_config(request_method="GET", renderer="dossier_gestion.html")
    def get(self) -> dict:
        url_data, lectures = get_all_urls(self.dossier)
        return {
            "dossier": self.dossier,
            "dossier_resource": self.context,
            "current_tab": "retrait",
            "team": self.dossier.team,
            "current_user": self.request.user,
            "url_data": url_data,
            "lectures": lectures,
        }

    @view_config(request_method="POST")
    def post(self) -> Response:
        lecture_str: str = self.request.POST.get("lecture")
        back_url = self.request.resource_url(self.context, "gestion")
        if not lecture_str.isdigit():
            self.request.session.flash(
                Message(
                    cls="warning", text="Erreur : impossible d'identifier la lecture."
                )
            )
            return HTTPFound(location=back_url)

        lecture = DBSession.query(Lecture).filter(Lecture.pk == int(lecture_str)).one()
        if not lecture:
            self.request.session.flash(
                Message(cls="warning", text="La lecture est introuvable.")
            )
            return HTTPFound(location=back_url)

        max_404: int = int(self.request.registry.settings.get("zam.lectures.max_404"))
        fetch_amendements(lecture_pk=lecture.pk, max_404=max_404)
        RefreshLecture(lecture=lecture, request=self.request)

        self.request.session.flash(
            Message(
                cls="success",
                text=f"La demande de mise à jour de la lecture {lecture} a été prise \
en compte avec un nombre d’écart max de {max_404} entre amendements.",
            )
        )
        return HTTPFound(location=back_url)


@view_defaults(context=DossierResource, name="retrait", permission="retrait")
class DossierRetraitForm(DossierViewBase):
    @view_config(request_method="POST")
    def post(self) -> Response:
        user_pk: Optional[int] = get_as_int_or_none(self.request.POST.get("pk"))
        if user_pk is None:
            self.request.session.flash(
                Message(cls="error", text=("Erreur lors du traitement de l'action."),)
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        user = DBSession.query(User).filter(User.pk == user_pk).one()

        if user.is_admin:
            self.request.session.flash(
                Message(cls="error", text=("Impossible de retirer un administrateur."),)
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        target = str(user.email)

        if self.request.user == user:
            self.request.session.flash(
                Message(
                    cls="warning",
                    text=("Impossible de se retirer de l’équipe soi-même."),
                )
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        if self.dossier.team.coordinators == [user]:
            self.request.session.flash(
                Message(
                    cls="warning",
                    text=(
                        "Un dossier législatif doit comporter au minimum \
un coordinateur."
                    ),
                )
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        self.dossier.team.remove_member(user)

        DossierRetrait.create(dossier=self.dossier, email=target, request=self.request)
        self.request.session.flash(
            Message(
                cls="success", text=(f"{target} a été retiré du dossier avec succès.")
            )
        )
        return HTTPFound(location=self.request.resource_url(self.context, "gestion"))


@view_defaults(
    context=DossierResource, name="set_contributeur", permission="set_contributeur"
)
class DossierSetContributeurForm(DossierViewBase):
    @view_config(request_method="POST")
    def post(self) -> Response:
        user_pk: Optional[int] = get_as_int_or_none(self.request.POST.get("pk"))
        if user_pk is None:
            self.request.session.flash(
                Message(cls="error", text=("Erreur lors du traitement de l'action."),)
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        user = DBSession.query(User).filter(User.pk == user_pk).one()

        if self.dossier.team.coordinators == [user]:
            self.request.session.flash(
                Message(
                    cls="warning",
                    text=(
                        "Un dossier législatif doit comporter au minimum \
un coordinateur."
                    ),
                )
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        self.dossier.team.remove_coordinateur(user)

        PasseEnContributeur.create(
            dossier=self.dossier, email=f"{user.email}", request=self.request
        )
        self.request.session.flash(
            Message(
                cls="success",
                text=(f"{user} a été passé Contributeur du dossier avec succès."),
            )
        )
        return HTTPFound(location=self.request.resource_url(self.context, "gestion"))


@view_defaults(
    context=DossierResource, name="set_coordinateur", permission="set_coordinateur"
)
class DossierSetCoordinateurForm(DossierViewBase):
    @view_config(request_method="POST")
    def post(self) -> Response:
        user_pk: Optional[int] = get_as_int_or_none(self.request.POST.get("pk"))
        if user_pk is None:
            self.request.session.flash(
                Message(cls="error", text=("Erreur lors du traitement de l'action."),)
            )
            return HTTPFound(
                location=self.request.resource_url(self.context, "gestion")
            )

        user = DBSession.query(User).filter(User.pk == user_pk).one()

        self.dossier.team.add_coordinators([user])

        # get commons mail values
        mailer = get_mailer(self.request)
        url_dossier = self.request.resource_url(self.request.context)
        contact_mail_from = self.request.registry.settings["zam.contact_mail_from"]

        invitations_sent = send_coordinator_invitations(
            mailer,
            self.request.user,
            [user],
            contact_mail_from,
            self.dossier.titre,
            url_dossier,
            self.request.registry,
        )

        if not invitations_sent:
            cls = "warning"
            message = "Aucun email n’a été envoyé."
        else:
            cls = "success"
            message = "Un email a été envoyé."
        message += "<br><br>"
        message += f"{user} a été passé Coordinateur du dossier avec succès."

        PasseEnCoordinateur.create(
            dossier=self.dossier, email=f"{user.email}", request=self.request
        )
        self.request.session.flash(Message(cls=cls, text=(message)))
        return HTTPFound(location=self.request.resource_url(self.context, "gestion"))


@view_defaults(context=DossierResource, name="archive", permission="archive_dossier")
class DossierArchiveForm(DossierViewBase):
    @view_config(request_method="POST")
    def post(self) -> Response:

        self.dossier.team.active = not self.dossier.team.active

        cls = "success"
        if self.dossier.team.active:
            message = "Le dossier a été activé."
        else:
            message = "Le dossier a été archivé."

        ArchiverDossier.create(
            dossier=self.dossier, active=self.dossier.team.active, request=self.request
        )
        ArchiverDossiersList.create(
            dossier=self.dossier, active=self.dossier.team.active, request=self.request
        )

        self.request.session.flash(Message(cls=cls, text=(message)))
        return HTTPFound(location=self.request.resource_url(self.context, "gestion"))


def get_message_from_invitation_result(
    invitations_sent: int, existing_members: List[User], bad_emails: List[str]
) -> Tuple[str, str]:

    cls = "warning"
    if invitations_sent:
        if invitations_sent > 1:
            message = "Invitations envoyées avec succès."
        else:
            message = "Invitation envoyée avec succès."
        cls = "success"
    else:
        message = "Aucune invitation n’a été envoyée."

    if existing_members:
        existing_emails = [user.email for user in existing_members]
        message += "<br><br>"
        if len(existing_emails) > 1:
            message += (
                f"Les adresses courriel {enumeration(existing_emails)} "
                "avaient déjà été invitées au dossier précédemment."
            )
        else:
            message += (
                f"L’adresse courriel {existing_emails[0]} "
                "avait déjà été invitée au dossier précédemment."
            )

    if bad_emails:
        message += "<br><br>"
        if len(bad_emails) > 1:
            message += (
                f"Les adresses courriel {enumeration(bad_emails)} "
                "sont mal formées ou non autorisées et n’ont pas été invitées."
            )
        else:
            message += (
                f"L’adresse courriel {bad_emails[0]} "
                "est mal formée ou non autorisée et n’a pas été invitée."
            )
    return cls, message


def get_all_urls(dossier: Dossier) -> Tuple[str, List]:
    url_data: str = ""
    if dossier.an_id:
        url_data = get_url_dossiers(int(dossier.an_id[5:7]))
    elif dossier.senat_id:
        url_data = build_rss_url(dossier.senat_id)

    lectures = [
        {
            "pk": lec.pk,
            "chambre": lec.format_chambre(),
            "organe": lec.format_organe(),
            "phase": lec.format_num_lecture(),
            "num_texte": lec.texte.numero,
            "libelle_texte": lec.format_texte(),
            "is_discuss": lec.chambre == Chambre.SENAT
            and len(list(derouleur_urls_and_mission_refs(lec))),
            "links_discuss": (
                [url[:-5] for url, mission in derouleur_urls_and_mission_refs(lec)]
                if lec.chambre == Chambre.SENAT
                else ""
            ),
            "link_texte": (
                get_texte_url_senat(lec.texte)
                if lec.chambre == Chambre.SENAT
                else get_parseur(lec).url
            ),
            "link_derouleur": (
                _build_amendements_url(lec)
                if lec.chambre == Chambre.SENAT
                else get_url_an_derouleur(lec)
            ),
            "missions": (
                MissionSenat.get_url_missions_senat(lec)
                if lec.elligible_mission_senat
                else ""
            ),
        }
        for lec in sorted(dossier.lectures, reverse=True)
    ]
    return url_data, lectures


def get_url_an_derouleur(lecture: Lecture) -> str:
    organe = repository.get_opendata_organe(lecture.organe)
    if organe is None:
        return ""
    return build_pattern(lecture=lecture, organe=organe["libelleAbrev"])
