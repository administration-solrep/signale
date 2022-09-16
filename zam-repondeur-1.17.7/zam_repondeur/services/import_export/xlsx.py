from collections import Counter
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from zam_repondeur.models import Amendement, Team, User

from .spreadsheet import FIELDS, HEADERS, export_amendement_for_spreadsheet

DARK_BLUE = Color(rgb="00182848")
WHITE = Color(rgb="00FFFFFF")


def export_xlsx(filename: str, amendements: List[Amendement]) -> Counter:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_xslsx_header_row(ws)
    counter = _export_xlsx_data_rows(ws, sorted(amendements))
    wb.save(filename)
    return counter


def _write_xslsx_header_row(ws: Worksheet) -> None:
    for column, value in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)
        cell.font = Font(color=WHITE, sz=8)


def _export_xlsx_data_rows(ws: Worksheet, amendements: Iterable[Amendement]) -> Counter:
    counter = Counter({"amendements": 0})
    for amend in amendements:
        amend_dict = {
            FIELDS[k]: v for k, v in export_amendement_for_spreadsheet(amend).items()
        }
        for column, value in enumerate(HEADERS, 1):
            cell = ws.cell(row=counter["amendements"] + 2, column=column)
            cell.value = amend_dict[value]
            cell.font = Font(sz=8)
            if cell.data_type == "f":
                cell.data_type = "s"  # Prevent errors with lines started with "="
        counter["amendements"] += 1
    return counter


def get_user_list_workbook(team: Team, slug: str) -> Workbook:
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = f"Equipe {slug}"[0:30]
    ws.sheet_properties.tabColor = "1072BA"

    ws.append(("Nom", "Email", "Rôle"))
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)

    for user in team.users:
        ws.append((user.name, user.email, User.get_role(user, team)))
    return wb
